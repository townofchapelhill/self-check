""" Parse and aggregate the selfcheck Hourly Excel file and output csv """

from openpyxl import load_workbook
from file_util import select_filename
import datetime
import pathlib
import re
import csv
import os
import filename_secrets

production_datasets_path = pathlib.Path(filename_secrets.productionStaging)
selfcheck_data_path =  pathlib.Path(filename_secrets.selfcheckStatistics)

# select input/output filenames
try:
        search_string = 'All-Hourly-LastWeek-*'
        self_check_input_hourly = select_filename(selfcheck_data_path, search_string)
        print(f'Input file: {self_check_input_hourly}')
        # output file is tagged with the last modification date of the input file
        timestamp = datetime.datetime.fromtimestamp(self_check_input_hourly.stat().st_mtime)
        filepath = "selfcheck-Hourly.csv"
        self_check_output_hourly = production_datasets_path.joinpath(filepath)
        print(f'Output file: {self_check_output_hourly}')
except  Exception:
    raise("Unable to determine input/output files")
	
# Identify fields for CSV output (daily)
csv_header = ['Date','Time','CheckoutSuccess','CheckoutFail','RenewedSuccess','RenewedFail','UserLoginSuccess','UserLoginFail','PaymentSuccess','PaymentFailed','CoinboxEmptyCount', 'SuccessfulTransactions', 'FailedTransactions', 'TotalTransactions']
output_fields = [0, 1, 2, 8, 9, 10, 11, 13, 14, 15, 17, 18, 19]

# Library open hours
library_hours = {
      "Monday":   {"open": 8, "close": 20},
      "Tuesday":  {"open": 8, "close": 20},
      "Wednesday":{"open": 8, "close": 20},
      "Thursday": {"open": 8, "close": 20},
      "Friday":   {"open": 8, "close": 18},
      "Saturday": {"open": 9, "close": 18},
      "Sunday":   {"open": 9, "close": 18}
    }

# Calculate DayofWeek
def DayofWeek(datestring):
    dt = datetime.datetime.strptime(datestring, '%d %B %Y')
    return datetime.date.strftime(dt, '%A')

# Open the workbook and set up access
try:
    workbook  = load_workbook(filename = str(self_check_input_hourly), read_only=True)
except  Exception:
    raise("Unable to parse input file")

# create output file & write header row
with open(self_check_output_hourly, 'w') as output_file:
    csvwriter = csv.writer(output_file, dialect='excel')
    csvwriter.writerow(csv_header)

    # Each day has a separate worksheet - iterate
    sheet_names = workbook.sheetnames
    for current_sheet in range(0, len(sheet_names)):
        # set access to the current worksheet
        workbook.active = current_sheet
        worksheet = workbook.active
        worksheet_list = list(worksheet.values)
        sheet_date = None
        # Select each row aggregate with a time in the first field
        for row in range(0, len(worksheet_list)):
            # clear the row array
            csv_output=['null']*len(csv_header)
            # if it finds the date in a row, save it and skip to next row
            if sheet_date is None:
                match = re.match('^\d+\s\w+\s\d{4}', worksheet_list[row][0])
                if match:
                    sheet_date=match.group()
                    # Calculate the Day of Week - used for Open Hours selection
                    dayofweek=DayofWeek(sheet_date)
                    continue
            # insert the date as the first field in the row
            csv_output[0] = sheet_date
            # process each row and write to CSV
            if worksheet_list[row][0] == "Total":
                # Final line of spreadsheet
                break
            # Find the hourly summary rows
            match = re.match('^\d{2}:\d{2}', worksheet_list[row][0])
            if match:
                # lookup if it is within Library Open Hours
                row_time = int(worksheet_list[row][0][:2])
                if  row_time >= library_hours[dayofweek]['open'] and row_time <= library_hours[dayofweek]['close']:
                    for element in range(0, len(output_fields)):
                        # Build the output row
                        csv_output[element+1] = worksheet_list[row][output_fields[element]]
                    csvwriter.writerow(csv_output)

# close and terminate
output_file.close()

